from banco_dados.consulta_funcionarios import ConsultaFuncionarios
import openpyxl
from openpyxl.styles import PatternFill
from lista_util.listas_sem_tabelas import lista_util, dicionario_ocupacional, cabecalho
from openpyxl.styles import Font
import getpass


class Planilha:
    consulta = ConsultaFuncionarios()
    cadastro_funcionario = consulta.cadastro_funcionario()
    situacao = consulta.verificando_situacao_funcionario()
    filhos = consulta.consulta_filhos()
    departamentos = consulta.consulta_departamento()
    servico = consulta.consulta_servico()
    cargos = consulta.consulta_cargos()
    sindicato = consulta.consulta_sindicato()
    banco = consulta.consulta_banco()
    operadora = consulta.consulta_operadora()
    municipio = consulta.consulta_municipio()
    pais = consulta.consulta_pais()

    tipo_horario, cor_raca, grau_instrucao, tipo_conta, categoria, emissor, residencia, \
    deficiencia, sindicalizado = lista_util()

    wb = openpyxl.Workbook()
    ws = wb.active

    negrito = Font(bold=True)
    fundo_cor = PatternFill(start_color="0099CCFF", end_color="0099CCFF", fill_type="solid")

    ordem_paramentro = [departamentos, servico, cargos, sindicato, tipo_horario, banco, municipio, pais,
                        pais, pais, municipio, tipo_conta, cor_raca, grau_instrucao, categoria, emissor,
                        residencia, deficiencia, sindicalizado]
    lista_cadastro = []
    lista_geral_cadastro = []

    user = getpass.getuser()

    # criação de função para conectar planilhas e modificar números por nomes
    def trocar_id_nomes(self):
        for cadastro in self.cadastro_funcionario:
            lista_cadastro = list(cadastro)

            self.lista_geral_cadastro.append(lista_cadastro)

        posicao = 2
        for parametro in self.ordem_paramentro:
            # print(posicao)
            self.troca_nome(parametro, posicao)
            posicao += 1

        # print(self.lista_geral_cadastro[2073])

    # Criação de função para buscar nome baseado no cod
    def troca_nome(self, listagem_parametros, posicao):
        for cadastro in self.lista_geral_cadastro:

            for listagem in listagem_parametros:

                if cadastro[posicao] == listagem[0]:
                    cadastro[posicao] = listagem[1]

    # criação de função para conectar planilhas e modificar números por nomes
    def plano_saude(self):

        for funcionario in self.lista_geral_cadastro:
            cod_funcionario = funcionario[0]
            plano_saude = funcionario[21]

            if plano_saude == 1:
                info_plano = self.consulta.consulta_plano(cod_funcionario)

                for operadora in self.operadora:
                    id_operadora = operadora[0]
                    nome_operadora = operadora[1]

                    if info_plano[0][0] == id_operadora:
                        funcionario[21] = nome_operadora
                        funcionario.insert(22, info_plano[0][1])

            else:
                funcionario[21] = ''
                funcionario.insert(22, '')

    # criação de função para conectar planilhas e modificar números por nomes
    def exame_toxicologico(self):
        for cadastro in self.lista_geral_cadastro:
            codi_func = cadastro[0]

            toxicologico = self.consulta.consulta_toxicologico(codi_func)

            if len(toxicologico) == 0:
                cadastro.append('')
                cadastro.append('')

            else:
                if toxicologico[0][1] == 1:
                    cadastro.append(toxicologico[0][0])
                    cadastro.append('Admissional')
                else:
                    cadastro.append(toxicologico[0][0])
                    cadastro.append('Demissional')

        # print(self.lista_geral_cadastro[2218])
        # print(self.lista_geral_cadastro[0])

    # criação de função para conectar planilhas e modificar números por nomes
    def exame_ocupacional(self):
        for cadastro in self.lista_geral_cadastro:
            codi_func = cadastro[0]

            ocupacional = self.consulta.consulta_ocupacional(codi_func)

            if len(ocupacional) == 0:
                cadastro.append('')
                cadastro.append('')
                cadastro.append('')
                cadastro.append('')
            else:
                dicionario, resultado = dicionario_ocupacional()
                tipo = dicionario[ocupacional[0][0]]
                result = resultado[ocupacional[0][2]]
                cadastro.append(tipo)
                cadastro.append(ocupacional[0][1])
                cadastro.append(result)
                cadastro.append(ocupacional[0][3])
        # print(self.lista_geral_cadastro[362])

    # criação de função para definir situação de atividade

    def situacao_funcionario(self):
        for cadastro in self.lista_geral_cadastro:
            codi_funcionario = cadastro[0]
            for situacao in self.situacao:
                if codi_funcionario == situacao[0]:
                    cadastro.insert(2, situacao[1])
                    break

    def dias_exp(self):

        for cadastro in self.lista_geral_cadastro:

            if cadastro[29] is None or cadastro[30] is None:
                cadastro.insert(31, '')

            else:

                dif = (cadastro[30] - cadastro[29]).days
                dif_correto = dif + 1
                cadastro.insert(31, dif_correto)

    def escrevendo_planilha(self):

        titulos_cabecalho = cabecalho()

        for coluna in range(1, len(titulos_cabecalho) + 1):
            self.ws.cell(row=1, column=coluna).font = self.negrito
            self.ws.cell(row=1, column=coluna).fill = self.fundo_cor

            self.ws.cell(row=1, column=coluna).value = titulos_cabecalho[coluna - 1]

        for cadastro in self.lista_geral_cadastro:
            cadastro_tupla = tuple(cadastro)

            self.lista_cadastro.append(cadastro_tupla)

        for linha in self.lista_cadastro:
            self.ws.append(linha)

        # self.wb.save(rf"C:\Users\{self.user}\Documents\Funcionarios FCB\Cadastro_Funcionarios.xlsx")

    def ordenando_colunas(self):
        # Move o banco
        self.ws.insert_cols(42)
        self.ws.move_range(f"I1:I{self.ws.max_row}", cols=33)
        self.ws.delete_cols(9, 1)

        # Move o CPF
        self.ws.insert_cols(4)
        self.ws.move_range(f"Y1:Y{self.ws.max_row}", cols=-21)
        self.ws.delete_cols(25, 1)

        # Move o PIS
        self.ws.insert_cols(5)
        self.ws.move_range(f"Z1:Z{self.ws.max_row}", cols=-21)
        self.ws.delete_cols(26, 1)

        # Move o Serviço
        self.ws.insert_cols(6)
        self.ws.move_range(f"H1:H{self.ws.max_row}", cols=-2)
        self.ws.delete_cols(8, 1)

        # Move o Cargo
        self.ws.insert_cols(7)
        self.ws.move_range(f"I1:I{self.ws.max_row}", cols=-2)
        self.ws.delete_cols(9, 1)

        # Move a Data de admissão
        self.ws.insert_cols(10)
        self.ws.move_range(f"AA1:AA{self.ws.max_row}", cols=-17)
        self.ws.delete_cols(27, 1)

        # Move o Vencimento das Férias
        self.ws.insert_cols(11)
        self.ws.move_range(f"AB1:AB{self.ws.max_row}", cols=-17)
        self.ws.delete_cols(28, 1)

        # Move a Categoria
        self.ws.insert_cols(12)
        self.ws.move_range(f"V1:V{self.ws.max_row}", cols=-10)
        self.ws.delete_cols(22, 1)

        # Move o salário
        self.ws.insert_cols(13)
        self.ws.move_range(f"AC1:AC{self.ws.max_row}", cols=-16)
        self.ws.delete_cols(29, 1)

        # Move dados da Experiência
        self.ws.insert_cols(14)
        self.ws.insert_cols(14)
        self.ws.insert_cols(14)
        self.ws.insert_cols(14)

        self.ws.move_range(f"AG1:AG{self.ws.max_row}", cols=-19)
        self.ws.move_range(f"AH1:AH{self.ws.max_row}", cols=-19)
        self.ws.move_range(f"AI1:AI{self.ws.max_row}", cols=-19)
        self.ws.move_range(f"AJ1:AJ{self.ws.max_row}", cols=-19)

        self.ws.delete_cols(33, 4)

        # Move dados da CTPS
        self.ws.insert_cols(18)
        self.ws.insert_cols(18)
        self.ws.insert_cols(18)
        self.ws.insert_cols(18)

        self.ws.move_range(f"AK1:AK{self.ws.max_row}", cols=-19)
        self.ws.move_range(f"AL1:AL{self.ws.max_row}", cols=-19)
        self.ws.move_range(f"AM1:AM{self.ws.max_row}", cols=-19)
        self.ws.move_range(f"AN1:AN{self.ws.max_row}", cols=-19)

        self.ws.delete_cols(37, 4)

        # Move dados do Horário
        self.ws.insert_cols(23)
        self.ws.insert_cols(23)
        self.ws.insert_cols(23)

        self.ws.move_range(f"AO1:AO{self.ws.max_row}", cols=-18)
        self.ws.move_range(f"AP1:AP{self.ws.max_row}", cols=-18)
        self.ws.move_range(f"AQ1:AQ{self.ws.max_row}", cols=-18)

        self.ws.delete_cols(41, 3)

        # Move dados do Banco
        self.ws.insert_cols(26, 3)

        self.ws.move_range(f"AS1:AS{self.ws.max_row}", cols=-19)
        self.ws.move_range(f"AR1:AR{self.ws.max_row}", cols=-17)
        self.ws.move_range(f"AT1:AT{self.ws.max_row}", cols=-17)

        self.ws.delete_cols(44, 3)

        self.wb.save(rf"C:\Users\{self.user}\Documents\Funcionarios FCB\Cadastro_Funcionarios.xlsx")
